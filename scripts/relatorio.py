"""
relatorio.py
Projeto: Monitoramento de Rastreamento Oncológico — AP 21
Lê o banco SQLite, envia para o BigQuery e gera Excel local.
"""

import pandas as pd
import sqlite3
import os
from pathlib import Path
from dotenv import load_dotenv
from datetime import datetime
from google.cloud import bigquery
from google.oauth2 import service_account

# =============================================================================
# CONFIGURAÇÕES
# =============================================================================

ROOT = Path(__file__).resolve().parents[1]
load_dotenv(ROOT / ".env")
DB_PATH = str(ROOT / "banco" / "rastreamento.db")

PROJETO_BQ = os.getenv("BIGQUERY_PROJECT", "rastreamento-oncologico-ap21")
DATASET_BQ = os.getenv("BIGQUERY_DATASET", "rastreamento_oncologico")
TABELA_BQ = os.getenv("BIGQUERY_TABLE", "populacao_alvo")
TABELA_COMPLETA = f"{PROJETO_BQ}.{DATASET_BQ}.{TABELA_BQ}"

SAIDA_DIR = str(ROOT / "saida" / "excel")
TIMESTAMP  = datetime.now().strftime("%Y%m%d_%H%M")
SAIDA_PATH = os.path.join(SAIDA_DIR, f"rastreamento_oncologico_{TIMESTAMP}.xlsx")

NOMES_PROGRAMA = {
    "mamografia":     "Mamografia",
    "citopatologico": "Citopatologico",
    "colonoscopia":   "Colonoscopia",
    "sangue_oculto":  "Sangue_Oculto_Fezes",
}

# Nomes sem espaço — compatíveis com BigQuery e Looker Studio
COLUNAS = {
    "cns":                  "CNS",
    "nome":                 "Nome_Paciente",
    "data_nascimento":      "Data_Nascimento",
    "idade":                "Idade",
    "sexo":                 "Sexo",
    "unidade":              "Unidade_VitaCare",
    "equipe":               "Equipe",
    "microarea":            "Microarea",
    "situacao_usuario":     "Situacao_Usuario",
    "programa_monitorado":  "Programa",
    "situacao":             "Situacao_SISREG",
    "status_rastreamento":  "Status_Rastreamento",
    "risco":                "Risco",
    "data_agendamento":     "Data_Agendamento",
    "data_solicitacao":     "Data_Solicitacao",
    "unidade_solicitante":  "Unidade_Solicitante_SISREG",
    "procedimento":         "Procedimento",
}

# Nomes amigáveis para o Excel
COLUNAS_EXCEL = {
    "CNS":                        "CNS",
    "Nome_Paciente":              "Nome do Paciente",
    "Data_Nascimento":            "Data de Nascimento",
    "Idade":                      "Idade",
    "Sexo":                       "Sexo",
    "Unidade_VitaCare":           "Unidade (VitaCare)",
    "Equipe":                     "Equipe",
    "Microarea":                  "Microárea",
    "Situacao_Usuario":           "Situação do Usuário",
    "Programa":                   "Programa",
    "Situacao_SISREG":            "Situação SISREG",
    "Status_Rastreamento":        "Status Rastreamento",
    "Risco":                      "Risco",
    "Data_Agendamento":           "Data Agendamento",
    "Data_Solicitacao":           "Data Solicitação",
    "Unidade_Solicitante_SISREG": "Unidade Solicitante (SISREG)",
    "Procedimento":               "Procedimento",
}

# =============================================================================
# LEITURA DO BANCO
# =============================================================================

def ler_banco():
    """Lê a tabela populacao_alvo do banco SQLite."""
    print(f"📂 Lendo banco: {DB_PATH}")
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql("SELECT * FROM populacao_alvo", conn)
    conn.close()
    print(f"   ✅ {len(df)} registros carregados.")
    return df


# =============================================================================
# PREPARAÇÃO DOS DADOS
# =============================================================================

def preparar_dados(df):
    """Organiza e formata os dados."""
    print("\n🔧 Preparando dados...")

    df["programa_monitorado"] = df["programa_monitorado"].map(NOMES_PROGRAMA).fillna(df["programa_monitorado"])

    for col in COLUNAS.keys():
        if col not in df.columns:
            df[col] = None

    df = df[list(COLUNAS.keys())].copy()
    df = df.rename(columns=COLUNAS)

    df = df.sort_values(
        by=["Unidade_VitaCare", "Programa", "Nome_Paciente"],
        ascending=True,
        na_position="last"
    )

    print(f"   ✅ Dados preparados: {len(df)} registros.")
    return df


# =============================================================================
# ENVIO PARA O BIGQUERY
# =============================================================================

def enviar_bigquery(df):
    """Envia o DataFrame para o BigQuery substituindo a tabela."""
    print(f"\n☁️  Enviando para o BigQuery...")
    print(f"   Tabela: {TABELA_COMPLETA}")

    df_bq = df.copy()
    df_bq["Idade"]            = pd.to_numeric(df_bq["Idade"], errors="coerce").astype("Int64")
    df_bq["Data_Nascimento"]  = pd.to_datetime(df_bq["Data_Nascimento"],  dayfirst=True, errors="coerce").dt.date
    df_bq["Data_Agendamento"] = pd.to_datetime(df_bq["Data_Agendamento"], dayfirst=True, errors="coerce").dt.date
    df_bq["Data_Solicitacao"] = pd.to_datetime(df_bq["Data_Solicitacao"], dayfirst=True, errors="coerce").dt.date

    cred_path = os.getenv("GOOGLE_APPLICATION_CREDENTIALS")
    if cred_path:
        credentials = service_account.Credentials.from_service_account_file(
            cred_path, scopes=["https://www.googleapis.com/auth/cloud-platform"]
        )
        client = bigquery.Client(credentials=credentials, project=PROJETO_BQ)
    else:
        # Usa Application Default Credentials (gcloud auth application-default login, Cloud Run etc.)
        client = bigquery.Client(project=PROJETO_BQ)

    job_config = bigquery.LoadJobConfig(
        write_disposition=bigquery.WriteDisposition.WRITE_TRUNCATE,
        autodetect=True,
    )

    job = client.load_table_from_dataframe(df_bq, TABELA_COMPLETA, job_config=job_config)
    job.result()

    tabela = client.get_table(TABELA_COMPLETA)
    print(f"   ✅ {tabela.num_rows} registros enviados com sucesso!")


# =============================================================================
# GERAÇÃO DO EXCEL
# =============================================================================

def gerar_excel(df):
    """Gera o arquivo Excel formatado com nomes amigáveis."""
    print(f"\n📊 Gerando Excel...")
    os.makedirs(SAIDA_DIR, exist_ok=True)

    df_excel = df.rename(columns=COLUNAS_EXCEL).copy()

    with pd.ExcelWriter(SAIDA_PATH, engine="openpyxl") as writer:
        df_excel.to_excel(writer, sheet_name="Rastreamento Oncológico", index=False)

        worksheet = writer.sheets["Rastreamento Oncológico"]

        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        cor_cabecalho = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        cor_linha_par = PatternFill(start_color="EEF3F8", end_color="EEF3F8", fill_type="solid")
        borda_fina    = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

        for cell in worksheet[1]:
            cell.fill      = cor_cabecalho
            cell.font      = Font(bold=True, color="FFFFFF", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = borda_fina

        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            fill = cor_linha_par if row_idx % 2 == 0 else None
            for cell in row:
                if fill:
                    cell.fill = fill
                cell.alignment = Alignment(vertical="center")
                cell.border    = borda_fina
                cell.font      = Font(size=9)

        larguras = {
            "CNS":                         18,
            "Nome do Paciente":            35,
            "Data de Nascimento":          16,
            "Idade":                        7,
            "Sexo":                         7,
            "Unidade (VitaCare)":          30,
            "Equipe":                      25,
            "Microárea":                   10,
            "Situação do Usuário":         20,
            "Programa":                    22,
            "Situação SISREG":             26,
            "Status Rastreamento":         20,
            "Risco":                       12,
            "Data Agendamento":            16,
            "Data Solicitação":            16,
            "Unidade Solicitante (SISREG)":30,
            "Procedimento":                30,
        }

        for col_idx, col_cells in enumerate(worksheet.columns, start=1):
            header  = col_cells[0].value
            largura = larguras.get(header, 15)
            worksheet.column_dimensions[get_column_letter(col_idx)].width = largura

        worksheet.freeze_panes             = "A2"
        worksheet.auto_filter.ref          = worksheet.dimensions
        worksheet.row_dimensions[1].height = 30

    print(f"   ✅ Excel gerado com sucesso!")
    print(f"   📁 {SAIDA_PATH}")


# =============================================================================
# EXECUÇÃO PRINCIPAL
# =============================================================================

if __name__ == "__main__":
    print("=" * 55)
    print("  Relatório — Rastreamento Oncológico AP 21")
    print("=" * 55)

    df = ler_banco()
    df = preparar_dados(df)
    enviar_bigquery(df)
    gerar_excel(df)

    print("\n✅ Relatório concluído!")
